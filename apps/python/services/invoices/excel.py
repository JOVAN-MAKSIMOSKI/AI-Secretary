"""Template fetch helpers for invoice and offer documents."""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Mapping, Optional
from uuid import UUID

from openpyxl import load_workbook

from services.invoices.calculations import apply_invoice_price_calculations
from services.invoices.client_lookup import fetch_client_contact_from_extracted
from services.langchain import run_invoice_extraction
from services.storage import supabase


# Centralized extraction value definitions.
EXTRACTED_USER_MESSAGE_KEYS: tuple[str, ...] = (
    "invoice_type",
    "value_date",
    "consignment_note_number",
    "order_number",
    "client_name",
    "description",
    "units",
    "price_per_unit",
    "tax_percentage",
)

DERIVED_EXTRACTION_KEYS: tuple[str, ...] = (
    "invoice_month",
    "invoice_year",
)

# Values resolved from businesses table.
BUSINESS_OUTPUT_KEYS: tuple[str, ...] = (
    "business",
)

# Values used by calculation service (services/invoices/calculations.py).
CALCULATION_INPUT_KEYS: tuple[str, ...] = (
    "invoice_type",
    "units",
    "price_per_unit",
    "tax_percentage",
)

CALCULATION_OUTPUT_KEYS: tuple[str, ...] = (
    "tax_percentage",
    "price_before_tax",
    "price_after_tax",
)

# Values used by client lookup service (services/invoices/client_lookup.py).
CLIENT_LOOKUP_INPUT_KEYS: tuple[str, ...] = (
    "client_name",
)

CLIENT_LOOKUP_OUTPUT_KEYS: tuple[str, ...] = (
    "client_id",
    "client_name",
    "client_tax_number",
    "client_email",
)

ALL_INVOICE_ENRICHMENT_KEYS: tuple[str, ...] = (
    *EXTRACTED_USER_MESSAGE_KEYS,
    *DERIVED_EXTRACTION_KEYS,
    *CALCULATION_OUTPUT_KEYS,
    *CLIENT_LOOKUP_OUTPUT_KEYS,
    *BUSINESS_OUTPUT_KEYS,
)


_ORDER_HINT_PATTERN = re.compile(r"\b(order|naracka|нарачка|narachka)\b", re.IGNORECASE)
_CONSIGNMENT_HINT_PATTERN = re.compile(
    r"\b(consignment|dispatch|ispratnica|испратница|товарен\s+лист)\b",
    re.IGNORECASE,
)
_ORDER_NUMBER_PATTERN = re.compile(
    r"\b(?:order(?:\s+number)?|naracka|нарачка|narachka)\s*[:#-]?\s*(\d+)\b",
    re.IGNORECASE,
)
_CONSIGNMENT_NUMBER_PATTERN = re.compile(
    r"\b(?:consignment(?:\s+note)?|dispatch(?:\s+note)?|ispratnica|испратница|товарен\s+лист)\s*[:#-]?\s*(\d+)\b",
    re.IGNORECASE,
)
_UNITS_PATTERN = re.compile(
    r"\b(\d+)\s*(?:units?|unit|pcs?|pieces?|edinici|единици|парчиња?)\b",
    re.IGNORECASE,
)
_PRICE_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(
        r"\b(?:price\s*(?:per\s*unit)?|unit\s*price|cena|цена)\s*[:=]?\s*([0-9][0-9.,]*)\b",
        re.IGNORECASE,
    ),
    re.compile(r"\b(?:at|po|по)\s*([0-9][0-9.,]*)\b", re.IGNORECASE),
)
_VALUE_DATE_PATTERN = re.compile(r"\b(\d{4}-\d{2}-\d{2})\b")


def _parse_number_like(value: Any) -> float | None:
    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        number = float(value)
        return number if number == number else None

    if not isinstance(value, str):
        return None

    token = re.sub(r"\s+", "", value.strip())
    if not re.match(r"^[+-]?\d[\d.,]*$", token):
        return None

    comma_count = token.count(",")
    dot_count = token.count(".")
    normalized = token

    if comma_count > 0 and dot_count > 0:
        if token.rfind(",") > token.rfind("."):
            normalized = token.replace(".", "").replace(",", ".")
        else:
            normalized = token.replace(",", "")
    elif comma_count > 0:
        last_comma_index = token.rfind(",")
        decimal_digits = len(token) - last_comma_index - 1
        if comma_count == 1 and decimal_digits <= 2:
            normalized = token.replace(",", ".")
        else:
            normalized = token.replace(",", "")
    elif dot_count > 0:
        last_dot_index = token.rfind(".")
        decimal_digits = len(token) - last_dot_index - 1
        if dot_count == 1 and decimal_digits <= 2:
            normalized = token
        else:
            normalized = token.replace(".", "")

    try:
        parsed = float(normalized)
    except ValueError:
        return None

    return parsed if parsed == parsed else None


def _first_pattern_match(pattern: re.Pattern[str], value: str) -> str | None:
    match = pattern.search(value)
    if not match:
        return None
    return match.group(1)


def _apply_rule_based_extraction_fallback(
    message: str,
    extracted: Mapping[str, Any],
) -> dict[str, Any]:
    """Backfill common invoice fields when LLM output is missing or misrouted."""
    result = dict(extracted)
    raw_message = message.strip()
    lowered_message = raw_message.lower()

    has_order_hint = _ORDER_HINT_PATTERN.search(raw_message) is not None
    has_consignment_hint = _CONSIGNMENT_HINT_PATTERN.search(raw_message) is not None

    consignment_raw = _first_pattern_match(_CONSIGNMENT_NUMBER_PATTERN, raw_message)
    order_raw = _first_pattern_match(_ORDER_NUMBER_PATTERN, raw_message)

    if "consignment_note_number" not in result and consignment_raw is not None:
        parsed_consignment = _parse_number_like(consignment_raw)
        if parsed_consignment is not None and parsed_consignment.is_integer():
            result["consignment_note_number"] = int(parsed_consignment)

    if "order_number" not in result and order_raw is not None:
        parsed_order = _parse_number_like(order_raw)
        if parsed_order is not None and parsed_order.is_integer():
            result["order_number"] = int(parsed_order)

    # Fix misrouted order/consignment number when only one hint is present in message text.
    if has_order_hint and not has_consignment_hint and "order_number" not in result:
        possible = _parse_number_like(result.get("consignment_note_number"))
        if possible is not None and possible.is_integer():
            result["order_number"] = int(possible)
            result.pop("consignment_note_number", None)

    if has_consignment_hint and not has_order_hint and "consignment_note_number" not in result:
        possible = _parse_number_like(result.get("order_number"))
        if possible is not None and possible.is_integer():
            result["consignment_note_number"] = int(possible)
            result.pop("order_number", None)

    if has_order_hint and not has_consignment_hint and "order_number" in result:
        result.pop("consignment_note_number", None)

    if has_consignment_hint and not has_order_hint and "consignment_note_number" in result:
        result.pop("order_number", None)

    if "units" not in result:
        units_raw = _first_pattern_match(_UNITS_PATTERN, raw_message)
        parsed_units = _parse_number_like(units_raw) if units_raw is not None else None
        if parsed_units is not None and parsed_units.is_integer() and parsed_units >= 0:
            result["units"] = int(parsed_units)

    if "price_per_unit" not in result:
        for pattern in _PRICE_PATTERNS:
            price_raw = _first_pattern_match(pattern, raw_message)
            parsed_price = _parse_number_like(price_raw) if price_raw is not None else None
            if parsed_price is not None:
                result["price_per_unit"] = parsed_price
                break

    if "value_date" not in result:
        date_raw = _first_pattern_match(_VALUE_DATE_PATTERN, raw_message)
        if date_raw is not None:
            result["value_date"] = date_raw

    if "invoice_type" not in result:
        if "transport" in lowered_message or "превоз" in lowered_message:
            result["invoice_type"] = "transport"
        elif (
            "goods" in lowered_message
            or "стока" in lowered_message
            or "roba" in lowered_message
            or "invoice" in lowered_message
            or "faktura" in lowered_message
            or "фактура" in lowered_message
        ):
            result["invoice_type"] = "goods"

    return result


def _require_tenant_owner_auth_id(tenant_id: str) -> str:
    business = (
        supabase.table("businesses")
        .select("owner_auth_id")
        .eq("owner_auth_id", tenant_id)
        .limit(1)
        .execute()
    )

    if not business.data:
        raise ValueError(f"Tenant '{tenant_id}' not found.")

    return business.data[0]["owner_auth_id"]


def _fetch_template_payload(
    tenant_id: str,
    template_table: str,
    fallback_name: str,
) -> dict[str, Any]:
    owner_auth_id = _require_tenant_owner_auth_id(tenant_id)

    query = (
        supabase.table(template_table)
        .select("id, tenant_id, name, storagePath, created_at")
        .eq("tenant_id", owner_auth_id)
        .order("created_at", desc=True)
        .limit(1)
    )

    response = query.execute()
    rows = response.data or []

    if not rows:
        raise ValueError(f"No template found in '{template_table}' for tenant '{tenant_id}'.")

    row = rows[0]
    template_bytes = supabase.storage.from_("documents").download(row["storagePath"])

    if not isinstance(template_bytes, (bytes, bytearray)):
        template_bytes = bytes(template_bytes)

    return {
        "template_id": row["id"],
        "template_name": row.get("name") or fallback_name,
        "template_source": row.get("storagePath"),
        "template_bytes": bytes(template_bytes),
    }


def fetch_business_values(owner_auth_id: str) -> dict[str, Any]:
    """Resolve business details needed for invoice generation."""
    response = (
        supabase.table("businesses")
        .select("name, address, tax_number, transaction_account, depositor, invoice_counter")
        .eq("owner_auth_id", owner_auth_id)
        .limit(1)
        .execute()
    )
    rows = response.data or []

    if not rows:
        raise ValueError(f"Business for owner '{owner_auth_id}' not found.")

    row = rows[0]
    return {
        "business": {
            "name": row.get("name"),
            "address": row.get("address"),
            "taxnumber": row.get("tax_number"),
            "transactionaccount": row.get("transaction_account"),
            "depositor": row.get("depositor"),
            "invoice_counter": row.get("invoice_counter"),
        }
    }


def fetch_invoice_template_payload(tenant_id: str) -> dict[str, Any]:
    """Fetch the stored invoice template bytes and metadata for a tenant."""
    return _fetch_template_payload(
        tenant_id=tenant_id,
        template_table="templatesInvoice",
        fallback_name="invoice-template",
    )


def fetch_offer_template_payload(tenant_id: str) -> dict[str, Any]:
    """Fetch the stored offer template bytes and metadata for a tenant."""
    return _fetch_template_payload(
        tenant_id=tenant_id,
        template_table="templatesOffer",
        fallback_name="offer-template",
    )


def load_invoice_template_workbook(tenant_id: str):
    """Load the stored invoice template into an openpyxl workbook."""
    payload = fetch_invoice_template_payload(tenant_id)
    return load_workbook(BytesIO(payload["template_bytes"]))


def _parse_extracted_date(raw_value: Any) -> date | None:
    if not isinstance(raw_value, str):
        return None

    value = raw_value.strip()
    if not value:
        return None

    try:
        # Accept YYYY-MM-DD and common ISO datetime values.
        return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _format_placeholder_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _normalize_invoice_period_fields(extracted: Mapping[str, Any]) -> dict[str, Any]:
    """Normalize invoice_month/year when models return combined or slash-prefixed tokens.

    Examples handled:
    - invoice_month: "05/2026" -> 5
    - invoice_month: "  /2026" -> removed (cannot infer month)
    - invoice_year: "/2026" -> 2026
    """
    result = dict(extracted)

    raw_month = result.get("invoice_month")
    month_value: int | None = None
    if isinstance(raw_month, int):
        month_value = raw_month
    elif isinstance(raw_month, float) and raw_month.is_integer():
        month_value = int(raw_month)
    elif isinstance(raw_month, str):
        token = raw_month.strip()
        if token:
            month_match = re.search(r"(?<!\d)(\d{1,2})(?:\s*/\s*\d{2,4})?", token)
            if month_match:
                month_value = int(month_match.group(1))

    if month_value is None or month_value < 1 or month_value > 12:
        result.pop("invoice_month", None)
    else:
        result["invoice_month"] = month_value

    raw_year = result.get("invoice_year")
    year_value: int | None = None
    if isinstance(raw_year, int):
        year_value = raw_year
    elif isinstance(raw_year, float) and raw_year.is_integer():
        year_value = int(raw_year)
    elif isinstance(raw_year, str):
        token = raw_year.strip()
        if token:
            year_match = re.search(r"(\d{4})", token)
            if year_match:
                year_value = int(year_match.group(1))

    if year_value is None or year_value < 1900 or year_value > 3000:
        result.pop("invoice_year", None)
    else:
        result["invoice_year"] = year_value

    return result


def _flatten_values(values: Mapping[str, Any], prefix: str = "") -> dict[str, str]:
    flattened: dict[str, str] = {}
    for key, value in values.items():
        current_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            flattened.update(_flatten_values(value, prefix=current_key))
            continue
        flattened[current_key] = _format_placeholder_value(value)
    return flattened


def _add_alias(replacements: dict[str, str], source_key: str, alias_key: str) -> None:
    if source_key in replacements and alias_key not in replacements:
        replacements[alias_key] = replacements[source_key]


def _add_template_aliases(replacements: dict[str, str]) -> dict[str, str]:
    # Derive plain counter token from invoice_number when value uses counter/year format.
    if "invoice_number" in replacements and "invoice_counter" not in replacements:
        invoice_number_value = replacements.get("invoice_number", "")
        if "/" in invoice_number_value:
            replacements["invoice_counter"] = invoice_number_value.split("/", 1)[0].strip()

    # Business aliases used in the uploaded template.
    _add_alias(replacements, "business.name", "Business.name")
    _add_alias(replacements, "business.address", "Business.address")
    _add_alias(replacements, "business.taxnumber", "Business.taxnumber")
    _add_alias(replacements, "business.transactionaccount", "Business.transactionaccount")
    _add_alias(replacements, "business.depositor", "Business.depositor")
    _add_alias(replacements, "business.invoice_counter", "Business.invoice_counter")

    # Client aliases from extracted + looked-up fields.
    _add_alias(replacements, "client.name", "Client.name")
    _add_alias(replacements, "client.address", "Client.address")
    _add_alias(replacements, "client.city", "Client.city")
    _add_alias(replacements, "client.taxnumber", "Client.taxnumber")
    _add_alias(replacements, "client_name", "Client.name")
    _add_alias(replacements, "client_tax_number", "Client.taxnumber")

    # Invoice aliases mapped to current payload field names.
    _add_alias(replacements, "invoice_number", "Invoice.number")
    _add_alias(replacements, "invoice_counter", "Invoice.counter")
    _add_alias(replacements, "invoice_month", "Invoice.month")
    _add_alias(replacements, "invoice_year", "Invoice.year")
    _add_alias(replacements, "invoice_date", "Invoice.invoice_date")
    _add_alias(replacements, "value_date", "Invoice.value_date")
    _add_alias(replacements, "consignment_note_number", "Invoice.consignment_note_number")
    _add_alias(replacements, "order_number", "Invoice.order_number")
    _add_alias(replacements, "description", "Invoice.description")
    _add_alias(replacements, "units", "Invoice.un")
    _add_alias(replacements, "price_per_unit", "Invoice.price")
    _add_alias(replacements, "price_before_tax", "Invoice.price_before_tax")
    _add_alias(replacements, "tax_percentage", "Invoice.tax_%")
    _add_alias(replacements, "tax_total", "Invoice.tax_total")
    _add_alias(replacements, "price_after_tax", "Invoice.price_after_tax")
    _add_alias(replacements, "price_after_tax_text", "Invoice.price_after_tax_text")

    # Lower-case invoice aliases for mixed style placeholders.
    _add_alias(replacements, "Invoice.month", "invoice.month")
    _add_alias(replacements, "Invoice.year", "invoice.year")
    _add_alias(replacements, "Invoice.counter", "invoice.counter")
    _add_alias(replacements, "Invoice.invoice_date", "invoice.invoice_date")

    return replacements


def render_invoice_template_bytes(
    template_bytes: bytes,
    values: Mapping[str, Any],
) -> bytes:
    """Replace placeholders in every workbook cell using provided values.

    Supported placeholder styles:
    - {{field}}
    - ${field}
    Nested mappings are available via dot paths, e.g. {{business.name}}.
    """
    workbook = load_workbook(BytesIO(template_bytes))
    replacements = _add_template_aliases(_flatten_values(values))
    direct_token_patterns = {
        key: re.compile(rf"(?<![A-Za-z0-9_.]){re.escape(key)}(?![A-Za-z0-9_.])")
        for key in replacements
    }

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                updated = cell.value
                for key, replacement in replacements.items():
                    updated = updated.replace(f"{{{{{key}}}}}", replacement)
                    updated = updated.replace(f"${{{key}}}", replacement)
                    # Also support direct token placeholders like Business.name and Invoice.number.
                    # Match whole tokens only to avoid collisions (e.g. Invoice.price_after_tax
                    # must not partially replace Invoice.price_after_tax_text).
                    updated = direct_token_patterns[key].sub(replacement, updated)

                if updated != cell.value:
                    cell.value = updated

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def extract_invoice_fields_from_message(
    message: str,
    owner_auth_id: str | None = None,
) -> dict[str, Any]:
    """Extract and enrich invoice fields from free text.

    This is the single orchestration point for invoice extraction-related logic:
    - user-message extraction (LLM)
    - deterministic price calculations
    - invoice month/year derivation from extracted date
    """
    extracted = run_invoice_extraction(
        message,
        allowed_keys=EXTRACTED_USER_MESSAGE_KEYS,
    )
    extracted = _apply_rule_based_extraction_fallback(message, extracted)
    extracted = _normalize_invoice_period_fields(extracted)
    extracted = apply_invoice_price_calculations(extracted)

    extracted_date = _parse_extracted_date(extracted.get("value_date"))
    if extracted_date is not None:
        extracted["invoice_month"] = extracted_date.month
        extracted["invoice_year"] = extracted_date.year

    if owner_auth_id:
        if isinstance(extracted.get("client_name"), str):
            try:
                extracted.update(
                    fetch_client_contact_from_extracted(
                        owner_auth_id,
                        extracted,
                    )
                )
            except ValueError:
                # Keep extraction response usable even when client lookup cannot resolve a match.
                pass
        extracted.update(fetch_business_values(owner_auth_id))

    return extracted
